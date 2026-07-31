from openpyxl import load_workbook

wb = load_workbook("ss.xlsx")
sheets = wb.active

for row in sheets.iter_rows(values_only=True):
    print(row)

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "mysheet"

ws.append(["name","email","city"])
ws.append(["mohsen","mohsen@m.co","thran"])
ws.append(["ali","a@A.c","z"])

wb.save("new_s.xlsx")